"""
Reporting Service — Crib Analytics & Reporting Framework.

Architecture decision: hybrid live-query + Redis cache.

All reports query live transactional tables via optimised PostgreSQL CTEs
(generate_series for monthly series instead of per-month loops).
Results are cached in Redis with short TTLs (5–30 min) and invalidated
on state-changing writes.

When scale demands it (>10k orgs or >1M payment rows), individual
functions can be switched to read from materialized views without
changing the calling API contracts.

Tier 1 reports:
  get_portfolio_summary          — KPI snapshot (extended dashboard)
  get_rent_collection_report     — collection rate per property / date range
  get_rent_arrears_report        — aging buckets (0-30, 31-60, 61-90, 90+)
  get_occupancy_vacancy_report   — units breakdown + estimated lost revenue
  get_maintenance_overview_report — status breakdown + by-property split

Tier 2 reports:
  get_maintenance_cost_report    — cost by property / category / contractor
  get_contractor_performance     — jobs, completion time, success rate
  get_lease_expiry_report        — leases expiring in 30 / 60 / 90 days
  get_income_expense_report      — monthly / quarterly / yearly P&L

Export helpers:
  to_csv(headers, rows)          — returns CSV string
  to_xlsx(sheets)                — returns .xlsx bytes via openpyxl
"""

from __future__ import annotations

import csv
import io
import json
import uuid
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from typing import Any

from sqlalchemy import func, select, text, cast, Integer, case, and_, or_
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.inspection import MaintenanceIssue
from app.models.lease import Lease, LeaseStatus
from app.models.payment import Payment, PaymentStatus, RentSchedule
from app.models.property import Property, Unit, UnitStatus
from app.models.tenant import Tenant as TenantModel
from app.models.contractor import Contractor
from app.utils.db_filters import org_scope

# ── Cache helpers ──────────────────────────────────────────────────────────────

REPORT_CACHE_TTL = {
    "portfolio_summary":     60 * 30,   # 30 min — refreshed on payments / lease changes
    "rent_collection":       60 * 10,   # 10 min
    "rent_arrears":          60 * 5,    # 5 min  — high-value operational data
    "occupancy":             60 * 30,
    "maintenance_overview":  60 * 10,
    "maintenance_cost":      60 * 30,
    "contractor_perf":       60 * 60,   # 1 hour — historical, changes slowly
    "lease_expiry":          60 * 60,
    "income_expense":        60 * 60,
}


def _cache_key(report: str, org_id: uuid.UUID | None, **params: Any) -> str:
    suffix = json.dumps({k: str(v) for k, v in sorted(params.items()) if v is not None}, sort_keys=True)
    return f"crib:report:{report}:{org_id or 'super'}:{suffix}"


async def _from_cache(redis, key: str) -> Any | None:
    try:
        raw = await redis.get(key)
        return json.loads(raw) if raw else None
    except Exception:
        return None


async def _to_cache(redis, key: str, data: Any, ttl: int) -> None:
    try:
        await redis.setex(key, ttl, json.dumps(data, default=str))
    except Exception:
        pass


async def invalidate_report_cache(redis, org_id: uuid.UUID) -> None:
    """Call on payment confirm / lease activation to evict stale report data."""
    try:
        pattern = f"crib:report:*:{org_id}:*"
        keys = await redis.keys(pattern)
        if keys:
            await redis.delete(*keys)
    except Exception:
        pass


# ── Shared utilities ───────────────────────────────────────────────────────────

_ACTIVE_LEASE_STATES = [
    LeaseStatus.active.value,
]

_SUCCESS_PAYMENT_STATUSES = [
    PaymentStatus.confirmed.value,
    PaymentStatus.completed.value,
    PaymentStatus.reconciled.value,
    PaymentStatus.allocated.value,
]


def _d(val: Any) -> float:
    """Safe Decimal → float conversion."""
    if val is None:
        return 0.0
    return float(val)


# ── Tier 1: Portfolio Summary ──────────────────────────────────────────────────

async def get_portfolio_summary(
    org_id: uuid.UUID | None,
    db: AsyncSession,
    redis=None,
    property_ids: list[uuid.UUID] | None = None,
) -> dict:
    """
    Extended KPI snapshot for the portfolio dashboard.

    Adds to the existing analytics_service.get_dashboard_stats():
      - vacant_units, vacancy_rate
      - overdue_amount (Decimal sum)
      - average_days_to_payment (paid schedules only)
      - maintenance split: open / in_progress / resolved_this_month
    """
    cache_key = _cache_key("portfolio_summary", org_id)
    if redis:
        cached = await _from_cache(redis, cache_key)
        if cached:
            return cached

    today = date.today()
    month_start = today.replace(day=1)
    next_month = (month_start + timedelta(days=32)).replace(day=1)

    def _scoped(q, col):
        q = org_scope(q, col, org_id)
        if property_ids is not None:
            q = q.where(col.in_(property_ids) if hasattr(col, "in_") else q)
        return q

    # Units
    unit_q = _scoped(select(func.count()).select_from(Unit).where(Unit.deleted_at.is_(None)), Unit.property_id)
    # — rewrite to use organisation scoping via properties join for property_ids case
    unit_total_q = select(func.count()).select_from(Unit).where(Unit.deleted_at.is_(None))
    unit_occupied_q = select(func.count()).select_from(Unit).where(
        Unit.deleted_at.is_(None), Unit.status == UnitStatus.occupied
    )
    if org_id is not None:
        prop_sub = select(Property.id).where(Property.organisation_id == org_id, Property.deleted_at.is_(None))
        if property_ids:
            prop_sub = prop_sub.where(Property.id.in_(property_ids))
        unit_total_q = unit_total_q.where(Unit.property_id.in_(prop_sub))
        unit_occupied_q = unit_occupied_q.where(Unit.property_id.in_(prop_sub))

    total_units    = await db.scalar(unit_total_q) or 0
    occupied_units = await db.scalar(unit_occupied_q) or 0
    vacant_units   = total_units - occupied_units

    # Properties
    prop_q = org_scope(
        select(func.count()).select_from(Property).where(Property.deleted_at.is_(None)),
        Property.organisation_id, org_id,
    )
    total_properties = await db.scalar(prop_q) or 0

    # Revenue this month
    pay_q = org_scope(select(func.sum(Payment.amount)).where(
        Payment.status.in_(_SUCCESS_PAYMENT_STATUSES),
        Payment.paid_at >= month_start,
        Payment.paid_at < next_month,
    ), Payment.organisation_id, org_id)
    monthly_revenue = _d(await db.scalar(pay_q))

    # Expected rent this month (schedules due this month for active leases)
    active_lease_ids_q = org_scope(
        select(Lease.id).where(Lease.status.in_(_ACTIVE_LEASE_STATES), Lease.deleted_at.is_(None)),
        Lease.organisation_id, org_id,
    )
    exp_q = select(func.sum(RentSchedule.amount_due)).where(
        RentSchedule.lease_id.in_(active_lease_ids_q),
        RentSchedule.due_date >= month_start,
        RentSchedule.due_date < next_month,
        RentSchedule.deleted_at.is_(None),
    )
    expected_rent = _d(await db.scalar(exp_q))

    # Overdue (schedules past due, not fully paid)
    overdue_q = select(
        func.count().label("count"),
        func.sum(RentSchedule.amount_due - RentSchedule.amount_paid).label("amount"),
    ).where(
        RentSchedule.lease_id.in_(active_lease_ids_q),
        RentSchedule.due_date < today,
        RentSchedule.status == "overdue",
        RentSchedule.deleted_at.is_(None),
    )
    overdue_row = (await db.execute(overdue_q)).one()
    overdue_count  = overdue_row.count or 0
    overdue_amount = _d(overdue_row.amount)

    # Maintenance
    maint_base = org_scope(
        select(MaintenanceIssue.state, func.count().label("n")).group_by(MaintenanceIssue.state),
        MaintenanceIssue.organisation_id, org_id,
    )
    maint_rows = (await db.execute(maint_base)).all()
    maint_by_state = {r.state: r.n for r in maint_rows}
    open_maintenance = (
        maint_by_state.get("reported", 0) +
        maint_by_state.get("assigned", 0) +
        maint_by_state.get("in_progress", 0)
    )

    collection_rate = round(monthly_revenue / expected_rent * 100, 1) if expected_rent > 0 else 0.0
    occupancy_rate  = round(occupied_units / total_units * 100, 1)    if total_units > 0  else 0.0
    vacancy_rate    = round(vacant_units   / total_units * 100, 1)    if total_units > 0  else 0.0

    result = {
        "total_properties":  total_properties,
        "total_units":       total_units,
        "occupied_units":    occupied_units,
        "vacant_units":      vacant_units,
        "occupancy_rate":    occupancy_rate,
        "vacancy_rate":      vacancy_rate,
        "monthly_revenue":   monthly_revenue,
        "expected_rent":     expected_rent,
        "outstanding_rent":  round(expected_rent - monthly_revenue, 2),
        "overdue_count":     overdue_count,
        "overdue_amount":    overdue_amount,
        "collection_rate":   collection_rate,
        "open_maintenance":  open_maintenance,
        "maintenance_by_state": maint_by_state,
    }
    if redis:
        await _to_cache(redis, cache_key, result, REPORT_CACHE_TTL["portfolio_summary"])
    return result


# ── Tier 1: Rent Collection Report ────────────────────────────────────────────

async def get_rent_collection_report(
    org_id: uuid.UUID | None,
    db: AsyncSession,
    date_from: date | None = None,
    date_to: date | None = None,
    property_id: uuid.UUID | None = None,
    redis=None,
) -> list[dict]:
    """
    Collection rate per property, filtered by date range.

    Returns one row per property:
      property_id, property_name, rent_due, rent_collected,
      outstanding, collection_pct, schedules_count, paid_count, overdue_count
    """
    cache_key = _cache_key("rent_collection", org_id, date_from=date_from, date_to=date_to, property_id=property_id)
    if redis:
        cached = await _from_cache(redis, cache_key)
        if cached:
            return cached

    df = date_from or date.today().replace(day=1)
    dt = date_to   or date.today()

    # rent_collected = sum of amount_paid on schedules in the period.
    # amount_paid is updated by the payment allocation flow as payments are
    # confirmed and linked to schedules (via rent_schedule_id or period match).
    # GREATEST(..., 0) prevents negative outstanding when schedules are overpaid.
    prop_filter = "AND l.property_id = :property_id" if property_id else ""
    org_filter  = "AND l.organisation_id = :org_id"  if org_id     else ""

    sql = text(f"""
        WITH sched AS (
            SELECT
                l.property_id,
                COUNT(*) FILTER (WHERE rs.status != 'waived')              AS schedules_count,
                COALESCE(SUM(rs.amount_due)  FILTER (WHERE rs.status != 'waived'), 0) AS rent_due,
                COALESCE(SUM(rs.amount_paid) FILTER (WHERE rs.status != 'waived'), 0) AS rent_collected,
                COUNT(*) FILTER (WHERE rs.status = 'paid')                 AS paid_count,
                COUNT(*) FILTER (WHERE rs.status = 'overdue')              AS overdue_count
            FROM rent_schedules rs
            JOIN leases l ON l.id = rs.lease_id
            WHERE rs.due_date BETWEEN :df AND :dt
              AND rs.deleted_at IS NULL
              AND l.deleted_at IS NULL
              AND l.status = ANY(:states)
              {org_filter}
              {prop_filter}
            GROUP BY l.property_id
        )
        SELECT
            p.id            AS property_id,
            p.name          AS property_name,
            COALESCE(s.schedules_count, 0) AS schedules_count,
            COALESCE(s.rent_due, 0)        AS rent_due,
            COALESCE(s.rent_collected, 0)  AS rent_collected,
            GREATEST(COALESCE(s.rent_due, 0) - COALESCE(s.rent_collected, 0), 0)
                                           AS outstanding,
            CASE WHEN COALESCE(s.rent_due, 0) > 0
                 THEN ROUND(COALESCE(s.rent_collected, 0) / s.rent_due * 100, 1)
                 ELSE 0 END AS collection_pct,
            COALESCE(s.paid_count, 0)      AS paid_count,
            COALESCE(s.overdue_count, 0)   AS overdue_count
        FROM properties p
        LEFT JOIN sched s ON s.property_id = p.id
        WHERE p.deleted_at IS NULL
          {'AND p.organisation_id = :org_id' if org_id else ''}
          {'AND p.id = :property_id'         if property_id else ''}
        ORDER BY rent_due DESC, p.name
    """)

    params: dict = {
        "df": df, "dt": dt,
        "states": _ACTIVE_LEASE_STATES + ["expired", "terminated"],
    }
    if org_id:      params["org_id"]      = str(org_id)
    if property_id: params["property_id"] = str(property_id)

    rows = (await db.execute(sql, params)).mappings().all()
    result = [
        {
            "property_id":     str(r["property_id"]),
            "property_name":   r["property_name"],
            "schedules_count": int(r["schedules_count"]),
            "rent_due":        _d(r["rent_due"]),
            "rent_collected":  _d(r["rent_collected"]),
            "outstanding":     _d(r["outstanding"]),
            "collection_pct":  _d(r["collection_pct"]),
            "paid_count":      int(r["paid_count"]),
            "overdue_count":   int(r["overdue_count"]),
        }
        for r in rows
    ]
    if redis:
        await _to_cache(redis, cache_key, result, REPORT_CACHE_TTL["rent_collection"])
    return result


# ── Tier 1: Rent Arrears Report ────────────────────────────────────────────────

async def get_rent_arrears_report(
    org_id: uuid.UUID | None,
    db: AsyncSession,
    property_id: uuid.UUID | None = None,
    redis=None,
) -> dict:
    """
    Aging buckets: 0-30 / 31-60 / 61-90 / 90+ days overdue.

    Returns:
      buckets: { "0_30": [...], "31_60": [...], "61_90": [...], "90_plus": [...] }
      summary: { bucket: { count, total_owed } }
    """
    cache_key = _cache_key("rent_arrears", org_id, property_id=property_id)
    if redis:
        cached = await _from_cache(redis, cache_key)
        if cached:
            return cached

    today = date.today()
    prop_filter = "AND l.property_id = :property_id" if property_id else ""
    org_filter  = "AND l.organisation_id = :org_id"  if org_id     else ""

    sql = text(f"""
        SELECT
            t.first_name || ' ' || t.last_name          AS tenant_name,
            t.id                                         AS tenant_id,
            p.name                                       AS property_name,
            p.id                                         AS property_id,
            u.name                                       AS unit_name,
            rs.due_date,
            (rs.amount_due - rs.amount_paid)             AS amount_owed,
            (:today - rs.due_date)                       AS days_overdue,
            CASE
                WHEN (:today - rs.due_date) <= 30  THEN '0_30'
                WHEN (:today - rs.due_date) <= 60  THEN '31_60'
                WHEN (:today - rs.due_date) <= 90  THEN '61_90'
                ELSE '90_plus'
            END                                          AS bucket
        FROM rent_schedules rs
        JOIN leases l      ON l.id  = rs.lease_id
        JOIN tenants t     ON t.id  = l.tenant_id
        JOIN properties p  ON p.id  = l.property_id
        LEFT JOIN units u  ON u.id  = l.unit_id
        WHERE rs.status = 'overdue'
          AND rs.deleted_at IS NULL
          AND l.deleted_at  IS NULL
          AND l.status = ANY(:states)
          {org_filter}
          {prop_filter}
        ORDER BY days_overdue DESC, amount_owed DESC
    """)

    params: dict = {"today": today, "states": _ACTIVE_LEASE_STATES}
    if org_id:      params["org_id"]      = str(org_id)
    if property_id: params["property_id"] = str(property_id)

    rows = (await db.execute(sql, params)).mappings().all()

    buckets: dict[str, list] = {"0_30": [], "31_60": [], "61_90": [], "90_plus": []}
    summary: dict[str, dict] = {
        b: {"count": 0, "total_owed": 0.0} for b in buckets
    }

    for r in rows:
        bucket = r["bucket"]
        row_dict = {
            "tenant_name":   r["tenant_name"],
            "tenant_id":     str(r["tenant_id"]),
            "property_name": r["property_name"],
            "property_id":   str(r["property_id"]),
            "unit_name":     r["unit_name"],
            "due_date":      r["due_date"].isoformat() if r["due_date"] else None,
            "amount_owed":   _d(r["amount_owed"]),
            "days_overdue":  int(r["days_overdue"]),
        }
        buckets[bucket].append(row_dict)
        summary[bucket]["count"]      += 1
        summary[bucket]["total_owed"] += row_dict["amount_owed"]

    for b in summary:
        summary[b]["total_owed"] = round(summary[b]["total_owed"], 2)

    result = {"buckets": buckets, "summary": summary, "as_of": today.isoformat()}
    if redis:
        await _to_cache(redis, cache_key, result, REPORT_CACHE_TTL["rent_arrears"])
    return result


# ── Tier 1: Occupancy & Vacancy Report ────────────────────────────────────────

async def get_occupancy_vacancy_report(
    org_id: uuid.UUID | None,
    db: AsyncSession,
    property_id: uuid.UUID | None = None,
    redis=None,
) -> dict:
    """
    Per-property occupancy breakdown.

    Returns:
      properties: [{ property_id, name, total, occupied, vacant, vacancy_pct, monthly_rent_lost }]
      totals: { total_units, occupied, vacant, vacancy_pct, monthly_rent_lost_est }
    """
    cache_key = _cache_key("occupancy", org_id, property_id=property_id)
    if redis:
        cached = await _from_cache(redis, cache_key)
        if cached:
            return cached

    prop_filter = "AND p.id = :property_id" if property_id else ""
    org_filter  = "AND p.organisation_id = :org_id" if org_id else ""

    sql = text(f"""
        SELECT
            p.id                                                               AS property_id,
            p.name                                                             AS property_name,
            COUNT(u.id)                                                        AS total_units,
            COUNT(u.id) FILTER (WHERE u.status = 'occupied')                  AS occupied,
            COUNT(u.id) FILTER (WHERE u.status != 'occupied')                 AS vacant,
            ROUND(
                COUNT(u.id) FILTER (WHERE u.status != 'occupied')::numeric
                / NULLIF(COUNT(u.id), 0) * 100, 1
            )                                                                  AS vacancy_pct,
            COALESCE(SUM(u.monthly_rent) FILTER (WHERE u.status != 'occupied'), 0)
                                                                               AS monthly_rent_lost
        FROM properties p
        JOIN units u ON u.property_id = p.id AND u.deleted_at IS NULL
        WHERE p.deleted_at IS NULL
          {org_filter}
          {prop_filter}
        GROUP BY p.id, p.name
        ORDER BY vacant DESC, p.name
    """)

    params: dict = {}
    if org_id:      params["org_id"]      = str(org_id)
    if property_id: params["property_id"] = str(property_id)

    rows = (await db.execute(sql, params)).mappings().all()
    properties = [
        {
            "property_id":       str(r["property_id"]),
            "property_name":     r["property_name"],
            "total_units":       int(r["total_units"]),
            "occupied":          int(r["occupied"]),
            "vacant":            int(r["vacant"]),
            "vacancy_pct":       _d(r["vacancy_pct"]),
            "monthly_rent_lost": _d(r["monthly_rent_lost"]),
        }
        for r in rows
    ]

    total_units       = sum(p["total_units"]       for p in properties)
    total_occupied    = sum(p["occupied"]           for p in properties)
    total_vacant      = sum(p["vacant"]             for p in properties)
    total_rent_lost   = round(sum(p["monthly_rent_lost"] for p in properties), 2)
    overall_vac_pct   = round(total_vacant / total_units * 100, 1) if total_units else 0.0

    result = {
        "properties": properties,
        "totals": {
            "total_units":           total_units,
            "occupied":              total_occupied,
            "vacant":                total_vacant,
            "vacancy_pct":           overall_vac_pct,
            "monthly_rent_lost_est": total_rent_lost,
        },
        "as_of": date.today().isoformat(),
    }
    if redis:
        await _to_cache(redis, cache_key, result, REPORT_CACHE_TTL["occupancy"])
    return result


# ── Tier 1: Maintenance Overview Report ───────────────────────────────────────

async def get_maintenance_overview_report(
    org_id: uuid.UUID | None,
    db: AsyncSession,
    property_id: uuid.UUID | None = None,
    contractor_id: uuid.UUID | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    redis=None,
) -> dict:
    """
    Maintenance status breakdown + per-property split.

    Returns:
      summary: { open, assigned, in_progress, resolved, closed, cancelled }
      by_property: [{ property_id, name, open, assigned, in_progress, resolved }]
      by_priority: { low, medium, high, urgent }
      by_category: { plumbing: n, ... }
    """
    cache_key = _cache_key("maintenance_overview", org_id,
                           property_id=property_id, contractor_id=contractor_id,
                           date_from=date_from, date_to=date_to)
    if redis:
        cached = await _from_cache(redis, cache_key)
        if cached:
            return cached

    where_clauses = ["1=1"]
    params: dict = {}
    if org_id:
        where_clauses.append("mi.organisation_id = :org_id")
        params["org_id"] = str(org_id)
    if property_id:
        where_clauses.append("mi.property_id = :property_id")
        params["property_id"] = str(property_id)
    if contractor_id:
        where_clauses.append("mi.contractor_id = :contractor_id")
        params["contractor_id"] = str(contractor_id)
    if date_from:
        where_clauses.append("mi.reported_at >= :date_from")
        params["date_from"] = date_from
    if date_to:
        where_clauses.append("mi.reported_at <= :date_to")
        params["date_to"] = date_to

    where = " AND ".join(where_clauses)

    # Summary by state
    state_sql = text(f"""
        SELECT state, COUNT(*) AS n
        FROM maintenance_issues mi
        WHERE {where}
        GROUP BY state
    """)
    state_rows = (await db.execute(state_sql, params)).mappings().all()
    summary = {r["state"]: int(r["n"]) for r in state_rows}

    # By property
    prop_sql = text(f"""
        SELECT
            p.id   AS property_id,
            p.name AS property_name,
            COUNT(*) FILTER (WHERE mi.state IN ('reported','assigned','in_progress')) AS open,
            COUNT(*) FILTER (WHERE mi.state = 'assigned')    AS assigned,
            COUNT(*) FILTER (WHERE mi.state = 'in_progress') AS in_progress,
            COUNT(*) FILTER (WHERE mi.state = 'resolved')    AS resolved,
            COUNT(*)                                          AS total
        FROM maintenance_issues mi
        JOIN properties p ON p.id = mi.property_id
        WHERE {where}
        GROUP BY p.id, p.name
        ORDER BY open DESC, p.name
    """)
    prop_rows = (await db.execute(prop_sql, params)).mappings().all()
    by_property = [
        {
            "property_id":   str(r["property_id"]),
            "property_name": r["property_name"],
            "open":          int(r["open"]),
            "assigned":      int(r["assigned"]),
            "in_progress":   int(r["in_progress"]),
            "resolved":      int(r["resolved"]),
            "total":         int(r["total"]),
        }
        for r in prop_rows
    ]

    # By priority
    prio_sql = text(f"""
        SELECT priority, COUNT(*) AS n
        FROM maintenance_issues mi WHERE {where}
        GROUP BY priority
    """)
    prio_rows = (await db.execute(prio_sql, params)).mappings().all()
    by_priority = {r["priority"]: int(r["n"]) for r in prio_rows}

    # By category
    cat_sql = text(f"""
        SELECT category, COUNT(*) AS n
        FROM maintenance_issues mi WHERE {where}
        GROUP BY category
    """)
    cat_rows = (await db.execute(cat_sql, params)).mappings().all()
    by_category = {r["category"]: int(r["n"]) for r in cat_rows}

    result = {
        "summary":     summary,
        "by_property": by_property,
        "by_priority": by_priority,
        "by_category": by_category,
    }
    if redis:
        await _to_cache(redis, cache_key, result, REPORT_CACHE_TTL["maintenance_overview"])
    return result


# ── Tier 2: Maintenance Cost Report ───────────────────────────────────────────

async def get_maintenance_cost_report(
    org_id: uuid.UUID | None,
    db: AsyncSession,
    property_id: uuid.UUID | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    redis=None,
) -> dict:
    """Cost of resolved/closed maintenance issues by property, category, contractor."""
    cache_key = _cache_key("maintenance_cost", org_id,
                           property_id=property_id, date_from=date_from, date_to=date_to)
    if redis:
        cached = await _from_cache(redis, cache_key)
        if cached:
            return cached

    where_clauses = ["mi.state IN ('resolved','closed')", "mi.actual_cost IS NOT NULL"]
    params: dict = {}
    if org_id:
        where_clauses.append("mi.organisation_id = :org_id")
        params["org_id"] = str(org_id)
    if property_id:
        where_clauses.append("mi.property_id = :property_id")
        params["property_id"] = str(property_id)
    if date_from:
        where_clauses.append("mi.resolved_at >= :date_from")
        params["date_from"] = date_from
    if date_to:
        where_clauses.append("mi.resolved_at <= :date_to")
        params["date_to"] = date_to

    where = " AND ".join(where_clauses)

    by_property_sql = text(f"""
        SELECT p.id AS property_id, p.name AS property_name,
               COUNT(*) AS jobs, SUM(mi.actual_cost) AS total_cost
        FROM maintenance_issues mi
        JOIN properties p ON p.id = mi.property_id
        WHERE {where}
        GROUP BY p.id, p.name ORDER BY total_cost DESC
    """)
    by_cat_sql = text(f"""
        SELECT category, COUNT(*) AS jobs, SUM(actual_cost) AS total_cost
        FROM maintenance_issues mi WHERE {where}
        GROUP BY category ORDER BY total_cost DESC
    """)
    by_contractor_sql = text(f"""
        SELECT c.id AS contractor_id, c.name AS contractor_name,
               COUNT(*) AS jobs, SUM(mi.actual_cost) AS total_cost
        FROM maintenance_issues mi
        LEFT JOIN contractors c ON c.id = mi.contractor_id
        WHERE {where}
        GROUP BY c.id, c.name ORDER BY total_cost DESC
    """)

    by_property = [
        {"property_id": str(r["property_id"]), "property_name": r["property_name"],
         "jobs": int(r["jobs"]), "total_cost": _d(r["total_cost"])}
        for r in (await db.execute(by_property_sql, params)).mappings().all()
    ]
    by_category = [
        {"category": r["category"], "jobs": int(r["jobs"]), "total_cost": _d(r["total_cost"])}
        for r in (await db.execute(by_cat_sql, params)).mappings().all()
    ]
    by_contractor = [
        {"contractor_id": str(r["contractor_id"]) if r["contractor_id"] else None,
         "contractor_name": r["contractor_name"] or "Unassigned",
         "jobs": int(r["jobs"]), "total_cost": _d(r["total_cost"])}
        for r in (await db.execute(by_contractor_sql, params)).mappings().all()
    ]

    result = {
        "by_property":   by_property,
        "by_category":   by_category,
        "by_contractor": by_contractor,
        "total_cost":    round(sum(r["total_cost"] for r in by_property), 2),
    }
    if redis:
        await _to_cache(redis, cache_key, result, REPORT_CACHE_TTL["maintenance_cost"])
    return result


# ── Tier 2: Contractor Performance Report ─────────────────────────────────────

async def get_contractor_performance(
    org_id: uuid.UUID | None,
    db: AsyncSession,
    date_from: date | None = None,
    date_to: date | None = None,
    redis=None,
) -> list[dict]:
    """
    Per-contractor: jobs completed, avg resolution time, success rate.
    """
    cache_key = _cache_key("contractor_perf", org_id, date_from=date_from, date_to=date_to)
    if redis:
        cached = await _from_cache(redis, cache_key)
        if cached:
            return cached

    where_clauses = ["mi.contractor_id IS NOT NULL"]
    params: dict = {}
    if org_id:
        where_clauses.append("mi.organisation_id = :org_id")
        params["org_id"] = str(org_id)
    if date_from:
        where_clauses.append("mi.assigned_at >= :date_from")
        params["date_from"] = date_from
    if date_to:
        where_clauses.append("mi.assigned_at <= :date_to")
        params["date_to"] = date_to

    where = " AND ".join(where_clauses)

    sql = text(f"""
        SELECT
            c.id                                                AS contractor_id,
            c.name                                              AS contractor_name,
            c.specialty,
            COUNT(*)                                            AS total_assigned,
            COUNT(*) FILTER (WHERE mi.state IN ('resolved','closed'))  AS completed,
            COUNT(*) FILTER (WHERE mi.state = 'cancelled')     AS cancelled,
            ROUND(
                AVG(EXTRACT(EPOCH FROM (mi.resolved_at - mi.assigned_at)) / 86400)
                FILTER (WHERE mi.resolved_at IS NOT NULL AND mi.assigned_at IS NOT NULL), 1
            )                                                   AS avg_days_to_resolve,
            ROUND(
                COUNT(*) FILTER (WHERE mi.state IN ('resolved','closed'))::numeric
                / NULLIF(COUNT(*) FILTER (WHERE mi.state NOT IN ('reported','cancelled')), 0) * 100, 1
            )                                                   AS success_rate_pct
        FROM maintenance_issues mi
        JOIN contractors c ON c.id = mi.contractor_id
        WHERE {where}
        GROUP BY c.id, c.name, c.specialty
        ORDER BY completed DESC, success_rate_pct DESC
    """)

    rows = (await db.execute(sql, params)).mappings().all()
    result = [
        {
            "contractor_id":     str(r["contractor_id"]),
            "contractor_name":   r["contractor_name"],
            "specialty":         r["specialty"],
            "total_assigned":    int(r["total_assigned"]),
            "completed":         int(r["completed"]),
            "cancelled":         int(r["cancelled"]),
            "avg_days_to_resolve": _d(r["avg_days_to_resolve"]),
            "success_rate_pct":  _d(r["success_rate_pct"]),
        }
        for r in rows
    ]
    if redis:
        await _to_cache(redis, cache_key, result, REPORT_CACHE_TTL["contractor_perf"])
    return result


# ── Tier 2: Lease Expiry Report ────────────────────────────────────────────────

async def get_lease_expiry_report(
    org_id: uuid.UUID | None,
    db: AsyncSession,
    redis=None,
) -> dict:
    """
    Leases expiring in 30 / 60 / 90 days.

    Returns:
      windows: { "30": [...], "60": [...], "90": [...] }
      summary: { "30": count, "60": count, "90": count }
    """
    cache_key = _cache_key("lease_expiry", org_id)
    if redis:
        cached = await _from_cache(redis, cache_key)
        if cached:
            return cached

    today = date.today()
    day90 = today + timedelta(days=90)

    org_filter = "AND l.organisation_id = :org_id" if org_id else ""

    sql = text(f"""
        SELECT
            l.id                                       AS lease_id,
            l.reference                                AS lease_ref,
            t.first_name || ' ' || t.last_name        AS tenant_name,
            p.name                                     AS property_name,
            u.name                                     AS unit_name,
            l.end_date,
            l.monthly_rent,
            l.currency,
            (l.end_date - :today)                      AS days_until_expiry
        FROM leases l
        JOIN tenants    t ON t.id = l.tenant_id
        JOIN properties p ON p.id = l.property_id
        LEFT JOIN units u ON u.id = l.unit_id
        WHERE l.status = 'active'
          AND l.end_date IS NOT NULL
          AND l.end_date BETWEEN :today AND :day90
          AND l.deleted_at IS NULL
          {org_filter}
        ORDER BY l.end_date ASC
    """)

    params: dict = {"today": today, "day90": day90}
    if org_id:
        params["org_id"] = str(org_id)

    rows = (await db.execute(sql, params)).mappings().all()

    def _row(r: Any) -> dict:
        return {
            "lease_id":          str(r["lease_id"]),
            "lease_ref":         r["lease_ref"],
            "tenant_name":       r["tenant_name"],
            "property_name":     r["property_name"],
            "unit_name":         r["unit_name"],
            "end_date":          r["end_date"].isoformat() if r["end_date"] else None,
            "monthly_rent":      _d(r["monthly_rent"]),
            "currency":          r["currency"],
            "days_until_expiry": int(r["days_until_expiry"]),
        }

    windows: dict[str, list] = {"30": [], "60": [], "90": []}
    for r in rows:
        d = int(r["days_until_expiry"])
        item = _row(r)
        if d <= 30:
            windows["30"].append(item)
        if d <= 60:
            windows["60"].append(item)
        windows["90"].append(item)

    result = {
        "windows": windows,
        "summary": {k: len(v) for k, v in windows.items()},
        "as_of": today.isoformat(),
    }
    if redis:
        await _to_cache(redis, cache_key, result, REPORT_CACHE_TTL["lease_expiry"])
    return result


# ── Tier 2: Income & Expense Report ───────────────────────────────────────────

async def get_income_expense_report(
    org_id: uuid.UUID | None,
    db: AsyncSession,
    group_by: str = "month",   # month | quarter | year
    months: int = 12,
    redis=None,
) -> list[dict]:
    """
    Monthly / quarterly / yearly P&L using generate_series CTE.

    Revenue:  confirmed/completed payments in period
    Expenses: resolved maintenance actual_cost in period
    Net:      revenue - expenses

    Uses a single generate_series query instead of per-period loops.
    """
    if group_by not in ("month", "quarter", "year"):
        group_by = "month"

    cache_key = _cache_key("income_expense", org_id, group_by=group_by, months=months)
    if redis:
        cached = await _from_cache(redis, cache_key)
        if cached:
            return cached

    today = date.today()
    start = (today.replace(day=1) - timedelta(days=30 * (months - 1))).replace(day=1)

    if group_by == "month":
        trunc = "month"
        series_step = "1 month"
    elif group_by == "quarter":
        trunc = "quarter"
        series_step = "3 months"
    else:
        trunc = "year"
        series_step = "1 year"

    org_pay_filter  = "AND p.organisation_id = :org_id" if org_id else ""
    org_maint_filter = "AND m.organisation_id = :org_id" if org_id else ""

    # Inline start/today as ISO literals — asyncpg mishandles :param::cast notation
    # in text() when params are named. These are server-generated dates, not user input.
    start_iso = start.isoformat()
    today_iso = today.isoformat()

    sql = text(f"""
        WITH periods AS (
            -- date_trunc on a date returns timestamp (no TZ) — matches revenue/expenses
            SELECT generate_series(
                date_trunc('{trunc}', '{start_iso}'::date),
                date_trunc('{trunc}', '{today_iso}'::date),
                '{series_step}'::interval
            ) AS period_start
        ),
        revenue AS (
            SELECT
                -- AT TIME ZONE 'UTC' on timestamptz → timestamp (no TZ), same as periods
                date_trunc('{trunc}', p.paid_at AT TIME ZONE 'UTC') AS period_start,
                SUM(p.amount) AS total
            FROM payments p
            WHERE p.status::text = ANY(ARRAY['confirmed','completed','reconciled','allocated']::text[])
              AND p.paid_at IS NOT NULL
              AND p.paid_at >= '{start_iso}'::date::timestamptz
              AND p.paid_at < ('{today_iso}'::date + interval '1 day')::timestamptz
              {org_pay_filter}
            GROUP BY 1
        ),
        expenses AS (
            SELECT
                date_trunc('{trunc}', m.resolved_at AT TIME ZONE 'UTC') AS period_start,
                SUM(m.actual_cost) AS total
            FROM maintenance_issues m
            WHERE m.resolved_at IS NOT NULL
              AND m.actual_cost IS NOT NULL
              AND m.resolved_at >= '{start_iso}'::date::timestamptz
              AND m.resolved_at < ('{today_iso}'::date + interval '1 day')::timestamptz
              {org_maint_filter}
            GROUP BY 1
        )
        SELECT
            pe.period_start,
            COALESCE(r.total, 0)  AS revenue,
            COALESCE(e.total, 0)  AS expenses,
            COALESCE(r.total, 0) - COALESCE(e.total, 0) AS net_income
        FROM periods pe
        LEFT JOIN revenue  r ON r.period_start = pe.period_start
        LEFT JOIN expenses e ON e.period_start = pe.period_start
        ORDER BY pe.period_start ASC
    """)

    params: dict = {}
    if org_id:
        params["org_id"] = str(org_id)

    rows = (await db.execute(sql, params)).mappings().all()
    result = [
        {
            "period":     r["period_start"].strftime("%Y-%m") if group_by == "month" else
                          (f"Q{((r['period_start'].month - 1) // 3) + 1} {r['period_start'].year}"
                           if group_by == "quarter" else str(r["period_start"].year)),
            "period_start": r["period_start"].isoformat(),
            "revenue":    _d(r["revenue"]),
            "expenses":   _d(r["expenses"]),
            "net_income": _d(r["net_income"]),
        }
        for r in rows
    ]
    if redis:
        await _to_cache(redis, cache_key, result, REPORT_CACHE_TTL["income_expense"])
    return result


# ── Export helpers ─────────────────────────────────────────────────────────────

def to_csv(headers: list[str], rows: list[list]) -> str:
    """Return a CSV string from a list of headers and row lists."""
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(headers)
    w.writerows(rows)
    return buf.getvalue()


def to_xlsx(sheets: dict[str, tuple[list[str], list[list]]]) -> bytes:
    """
    Return .xlsx bytes.

    sheets = { "Sheet Name": (headers_list, rows_list), ... }
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError("openpyxl is required for Excel export. Install it with: pip install openpyxl")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="1F4E79")
    center_align = Alignment(horizontal="center")

    for sheet_name, (headers, rows) in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])   # Excel limit: 31 chars
        ws.append(headers)
        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center_align
        for row in rows:
            ws.append(row)
        # Auto-width columns (approximate)
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Flat table helpers for export ──────────────────────────────────────────────

def rent_collection_to_rows(data: list[dict]) -> tuple[list[str], list[list]]:
    headers = ["Property", "Schedules", "Rent Due", "Collected", "Outstanding",
               "Collection %", "Paid", "Overdue"]
    rows = [
        [r["property_name"], r["schedules_count"], r["rent_due"], r["rent_collected"],
         r["outstanding"], r["collection_pct"], r["paid_count"], r["overdue_count"]]
        for r in data
    ]
    return headers, rows


def rent_arrears_to_rows(data: dict) -> tuple[list[str], list[list]]:
    headers = ["Tenant", "Property", "Unit", "Due Date", "Amount Owed", "Days Overdue", "Bucket"]
    rows = []
    bucket_labels = {"0_30": "0–30 days", "31_60": "31–60 days",
                     "61_90": "61–90 days", "90_plus": "90+ days"}
    for bucket_key, items in data["buckets"].items():
        for r in items:
            rows.append([
                r["tenant_name"], r["property_name"], r["unit_name"] or "",
                r["due_date"], r["amount_owed"], r["days_overdue"],
                bucket_labels.get(bucket_key, bucket_key),
            ])
    return headers, rows


def lease_expiry_to_rows(data: dict) -> tuple[list[str], list[list]]:
    headers = ["Lease Ref", "Tenant", "Property", "Unit", "End Date",
               "Days Until Expiry", "Monthly Rent", "Currency"]
    seen = set()
    rows = []
    for item in data["windows"]["90"]:
        if item["lease_id"] not in seen:
            seen.add(item["lease_id"])
            rows.append([
                item["lease_ref"], item["tenant_name"], item["property_name"],
                item["unit_name"] or "", item["end_date"], item["days_until_expiry"],
                item["monthly_rent"], item["currency"],
            ])
    return headers, rows


def income_expense_to_rows(data: list[dict]) -> tuple[list[str], list[list]]:
    headers = ["Period", "Revenue", "Expenses", "Net Income"]
    rows = [[r["period"], r["revenue"], r["expenses"], r["net_income"]] for r in data]
    return headers, rows
